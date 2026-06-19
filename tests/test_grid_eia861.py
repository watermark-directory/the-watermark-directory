"""EIA-861 per-utility retail connector (#94) + the PJM annual-load pull (EIA-930).

Two layers of coverage: (1) the xlsx reduction (`_reduce_sales_ult_cust`) is exercised
against an in-memory Sales-to-Ultimate-Customers workbook — so the parsing/summing logic
runs offline, not only on a live pull; (2) the public `fetch_utility_retail` /
`fetch_ba_annual_load` replay committed fixtures and assemble the provenance-tagged model.
Hermetic — no network.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

from bosc.config import Settings
from bosc.connectors import OfflineError
from bosc.grid.eia861 import (
    Eia861Error,
    _reduce_sales_ult_cust,
    _reduce_short_form,
    fetch_utility_retail,
)
from bosc.grid.interchange import fetch_ba_annual_load

REPO_ROOT = Path(__file__).resolve().parents[1]
FIXTURES = REPO_ROOT / "tests" / "fixtures" / "economics"


@pytest.fixture
def econ_settings() -> Settings:
    return Settings(data_dir=REPO_ROOT / "data", econ_offline=True, econ_fixtures_dir=FIXTURES)


def _sales_row(num: int, name: str, svc: str, state: str, rev: float, sales: float, cust: float):
    """A 24-column Sales_Ult_Cust data row (identity cols + zeroed sectors + TOTAL block)."""
    row = [None] * 24
    row[0], row[1], row[2] = 2024, num, name  # year, number, name
    row[3], row[4], row[6] = "A", svc, state  # part, service type, state
    row[21], row[22], row[23] = rev, sales, cust  # TOTAL revenues / sales / customers
    return row


def _sales_ult_cust_zip() -> bytes:
    """A minimal EIA-861 zip with a Sales_Ult_Cust_2024.xlsx ('States' sheet)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "States"
    ws.append(["header row 1"])  # rows 1-3 are headers; data starts at row 4
    ws.append(["header row 2"])
    ws.append(["Data Year", "Utility Number", "Utility Name"])
    # Ohio Power Co (#14006), OH: a Bundled (SSO) row + a Delivery (shopping) row.
    ws.append(_sales_row(14006, "Ohio Power Co", "Bundled", "OH", 1_463_723, 7_866_324, 611_084))
    ws.append(_sales_row(14006, "Ohio Power Co", "Delivery", "OH", 2_053_383, 40_786_583, 922_181))
    ws.append(
        _sales_row(14006, "Ohio Power Co", "Bundled", "WV", 999, 999, 9)
    )  # other state — ignored
    ws.append(
        _sales_row(55, "Some Other Util", "Bundled", "OH", 1, 1, 1)
    )  # other utility — ignored
    buf = io.BytesIO()
    wb.save(buf)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w") as z:
        z.writestr("Sales_Ult_Cust_2024.xlsx", buf.getvalue())
    return zbuf.getvalue()


def test_reduce_sales_ult_cust_sums_service_types() -> None:
    out = _reduce_sales_ult_cust(_sales_ult_cust_zip(), year=2024, utility_number=14006, state="OH")
    assert out["rows"] == 2  # only the two OH Ohio-Power rows
    # Total = Bundled + Delivery (all energy on the utility's wires).
    assert out["total_sales_mwh"] == pytest.approx(7_866_324 + 40_786_583)
    assert out["total_customers"] == pytest.approx(611_084 + 922_181)
    # Bundled-only revenue/sales (for the full-service average price).
    assert out["bundled_sales_mwh"] == pytest.approx(7_866_324)
    assert out["bundled_revenue_thousand_usd"] == pytest.approx(1_463_723)


def test_reduce_missing_utility_raises() -> None:
    with pytest.raises(Eia861Error):
        _reduce_sales_ult_cust(_sales_ult_cust_zip(), year=2024, utility_number=99999, state="OH")


def _short_form_row(num: int, name: str, ownership: str, state: str, rev, sales, cust):
    """A 9-column EIA-861 '861S' short-form row (year, number, name, ownership, state, BA, totals)."""
    row = [None] * 9
    row[0], row[1], row[2] = 2024, num, name
    row[3], row[4], row[5] = ownership, state, "PJM"
    row[6], row[7], row[8] = rev, sales, cust  # Total Revenue / Sales / Customers
    return row


def _short_form_zip(*, with_full_sheet: bool = False) -> bytes:
    """A minimal EIA-861 zip with a Short_Form_2024.xlsx ('861S' sheet); optionally a full sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "861S"
    ws.append(["header row 1"])  # data starts at row 2
    ws.append(
        _short_form_row(2439, "City of Bryan - (OH)", "Municipal", "OH", 17256.3, 160479, 5814)
    )
    ws.append(
        _short_form_row(14229, "City of Ottawa - (KS)", "Municipal", "KS", 1, 1, 1)
    )  # ignored
    buf = io.BytesIO()
    wb.save(buf)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w") as z:
        z.writestr("Short_Form_2024.xlsx", buf.getvalue())
        if with_full_sheet:
            # A full Sales sheet that does NOT contain #2439 — exercises the real fallback
            # (utility present in the bulk file, but only on the short form).
            z.writestr("Sales_Ult_Cust_2024.xlsx", _sales_ult_cust_zip_inner())
    return zbuf.getvalue()


def _sales_ult_cust_zip_inner() -> bytes:
    """The Sales_Ult_Cust workbook bytes (an IOU only — no municipal #2439)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "States"
    ws.append(["h1"])
    ws.append(["h2"])
    ws.append(["Data Year", "Utility Number", "Utility Name"])
    ws.append(_sales_row(14006, "Ohio Power Co", "Bundled", "OH", 1_463_723, 7_866_324, 611_084))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_reduce_short_form_reads_municipal_total() -> None:
    out = _reduce_short_form(_short_form_zip(), year=2024, utility_number=2439, state="OH")
    assert out["rows"] == 1
    assert out["form"] == "861s"
    assert out["ownership"] == "Municipal"
    assert out["utility_name"] == "City of Bryan - (OH)"
    assert out["total_sales_mwh"] == pytest.approx(160479)
    assert out["total_customers"] == pytest.approx(5814)
    # A municipal's whole retail is bundled full-service: total == bundled.
    assert out["bundled_sales_mwh"] == pytest.approx(160479)
    assert out["bundled_revenue_thousand_usd"] == pytest.approx(17256.3)


def test_reduce_short_form_missing_utility_raises() -> None:
    with pytest.raises(Eia861Error):
        _reduce_short_form(_short_form_zip(), year=2024, utility_number=99999, state="OH")


def test_fetch_utility_retail_falls_back_to_short_form(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A municipal absent from the full Sales sheet resolves via the EIA-861 short form."""
    import bosc.grid.eia861 as eia861

    monkeypatch.setattr(
        eia861, "_ensure_zip", lambda settings, year: _short_form_zip(with_full_sheet=True)
    )
    settings = Settings(data_dir=tmp_path, econ_offline=False)  # live path → fetch() runs
    up = fetch_utility_retail(utility_number=2439, state="OH", year=2024, settings=settings)
    assert "Bryan" in up.utility
    assert up.ownership == "Municipal"
    assert "861S Short Form" in up.eia_source or "861S" in up.eia_source
    assert up.retail_sales_gwh.value == pytest.approx(160.5, abs=0.1)  # 160,479 MWh
    assert up.customers is not None and up.customers.value == 5814
    # Full-service price = total revenue / total sales (no delivery-only split on the short form).
    assert up.avg_price_cents_kwh is not None
    assert up.avg_price_cents_kwh.value == pytest.approx(17256.3 / 160479 * 100.0, abs=0.05)
    assert "municipal" in (up.avg_price_cents_kwh.citation or "").lower()


def test_fetch_utility_retail_offline(econ_settings: Settings) -> None:
    up = fetch_utility_retail(settings=econ_settings)
    assert "AEP Ohio" in up.utility
    # ~48.65 TWh retail, ~1.53M customers, all connector-sourced.
    assert up.retail_sales_gwh.source == "connector" and up.retail_sales_gwh.verified
    assert 40_000 < up.retail_sales_gwh.value < 60_000
    assert up.customers is not None and up.customers.value > 1_000_000
    # Bundled full-service avg price (delivery-only excluded) — a sane residential-ish ¢/kWh.
    assert up.avg_price_cents_kwh is not None
    assert 10.0 < up.avg_price_cents_kwh.value < 30.0
    assert "delivery-only" in (up.avg_price_cents_kwh.citation or "")


def test_fetch_utility_retail_offline_miss_raises(tmp_path: Path) -> None:
    with pytest.raises(OfflineError):
        fetch_utility_retail(
            settings=Settings(data_dir=tmp_path, econ_offline=True, econ_fixtures_dir=None)
        )


def test_pjm_annual_load_offline(econ_settings: Settings) -> None:
    load = fetch_ba_annual_load(ba="PJM", year=2024, settings=econ_settings)
    assert load.source == "connector" and load.verified
    # PJM ~ 800 TWh/yr; sanity band in GWh.
    assert 700_000 < load.value < 950_000
    assert "EIA-930" in load.citation and "Eastern" in load.citation
