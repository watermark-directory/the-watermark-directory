"""EIA-861 per-utility retail profile — the serving utility's sales/customers/price (#94).

The serving utility's retail sales, customers, and average price come from the EIA-861
Annual Electric Power Industry Report "Sales to Ultimate Customers" file — the
**per-utility entity data the v2 ``seriesid`` route does not expose**. We download the
annual bulk zip, parse the one utility's rows from the ``States`` sheet, and reduce to a
tiny payload (so the cached payload / committed fixture stays small).

AEP Ohio (Ohio Power Co, EIA utility #14006) reports in a restructured-Ohio split: a
**Bundled** (standard-service-offer) row + a **Delivery** (shopping / delivery-only) row.
Total retail = their sum (all energy delivered on the utility's wires). The **average
price** is the *bundled* (full-service) revenue/sales — the delivery-only rows carry only
the wires charge (generation is paid to a competitive supplier), so a blended price would
understate the all-in cost.

Connector contract: pure ``fn(..., settings) -> pydantic`` via the shared ``cached_get``
(econ cache/offline/fixtures, shared with #91/#94/#95). The bulk zip downloads to
``econ_cache_dir/eia861/`` on the live path only; offline replays the committed fixture.
The EIA-861 bulk file needs no API key (public download).
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any, cast

import httpx
import openpyxl

from watermark.config import Settings, get_settings
from watermark.connectors import cached_get, to_float
from watermark.grid.model import UtilityProfile
from watermark.hydrology.model import ProvenancedValue
from watermark.logging import get_logger

log = get_logger(__name__)

# "Sales_Ult_Cust" column indices (1-based) on the "States" sheet; data starts at row 4
# (rows 1-3 are the section / sub / column header rows). Selected by NAME-verified
# position, never guessed: cols 2/3/5/7 identity, cols 22/23/24 the TOTAL block.
_COL_UTILNUM = 2  # Utility Number
_COL_NAME = 3  # Utility Name
_COL_SVC = 5  # Service Type ("Bundled" / "Energy" / "Delivery")
_COL_STATE = 7  # State
_COL_OWNERSHIP = 8  # Ownership ("Investor Owned" / "Municipal" / "Cooperative" / ...)
_COL_TOT_REV = 22  # TOTAL Revenues (Thousand Dollars)
_COL_TOT_SALES = 23  # TOTAL Sales (Megawatthours)
_COL_TOT_CUST = 24  # TOTAL Customers (Count)
_DATA_FIRST_ROW = 4

# EIA-861 Short Form ("861S" sheet) column indices (1-based); header is row 1, data from row 2.
# Small utilities — municipals, small cooperatives — file the short form and are ABSENT from the
# full Sales-to-Ultimate-Customers sheet, reporting a single annual TOTALS line here (no
# service-type split). Selected by NAME-verified position from the 861S header.
_SF_COL_UTILNUM = 2  # Utility Number
_SF_COL_NAME = 3  # Utility Name
_SF_COL_OWNERSHIP = 4  # Ownership
_SF_COL_STATE = 5  # State
_SF_COL_TOT_REV = 7  # Total Revenue (Thousand Dollars)
_SF_COL_TOT_SALES = 8  # Total Sales (MWh)
_SF_COL_TOT_CUST = 9  # Total Customers
_SF_DATA_FIRST_ROW = 2


class Eia861Error(RuntimeError):
    """The EIA-861 bulk file was missing the expected sheet/utility (format drift)."""


def _num(x: Any) -> float:
    return to_float(x, 0.0)


def _sales_ult_cust_name(year: int) -> str:
    return f"Sales_Ult_Cust_{year}.xlsx"


def _zip_path(settings: Settings, year: int) -> Path:
    return settings.econ_cache_dir / "eia861" / f"f861{year}.zip"


def _ensure_zip(settings: Settings, year: int) -> bytes:
    """Return the EIA-861 annual zip bytes, downloading to cache on first use."""
    path = _zip_path(settings, year)
    if path.is_file():
        return path.read_bytes()
    url = f"{settings.eia861_base_url}/f861{year}.zip"
    log.info("eia861.download", url=url, year=year)
    data = httpx.get(
        url, timeout=max(settings.econ_request_timeout_s, 120.0), follow_redirects=True
    ).content
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return data


def _reduce_sales_ult_cust(
    zip_bytes: bytes, *, year: int, utility_number: int, state: str
) -> dict[str, Any]:
    """Reduce the Sales to Ultimate Customers file to one utility's summed totals."""
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    want = _sales_ult_cust_name(year).replace(" ", "_").lower()
    matches = [n for n in z.namelist() if n.replace(" ", "_").lower() == want]
    if not matches:
        raise Eia861Error(f"{_sales_ult_cust_name(year)} not found in EIA-861 {year} zip")
    wb = openpyxl.load_workbook(io.BytesIO(z.read(matches[0])), read_only=True, data_only=True)
    ws = wb["States"]

    total_sales_mwh = 0.0
    total_customers = 0.0
    bundled_rev_thousand = 0.0
    bundled_sales_mwh = 0.0
    n_rows = 0
    util_name: str | None = None
    ownership: str = ""
    for row in ws.iter_rows(min_row=_DATA_FIRST_ROW, values_only=True):
        if not row:
            continue
        raw_num = row[_COL_UTILNUM - 1]
        if not isinstance(raw_num, (int, float, str)):
            continue
        try:
            num = int(float(raw_num))  # tolerate int / "14006" / 14006.0
        except (TypeError, ValueError):
            continue
        if num != utility_number or str(row[_COL_STATE - 1] or "").strip() != state:
            continue
        n_rows += 1
        util_name = str(row[_COL_NAME - 1] or "").strip() or util_name
        ownership = str(row[_COL_OWNERSHIP - 1] or "").strip() or ownership
        sales = _num(row[_COL_TOT_SALES - 1])
        total_sales_mwh += sales
        total_customers += _num(row[_COL_TOT_CUST - 1])
        if str(row[_COL_SVC - 1] or "").strip().lower() == "bundled":
            bundled_rev_thousand += _num(row[_COL_TOT_REV - 1])
            bundled_sales_mwh += sales
    wb.close()
    if n_rows == 0:
        raise Eia861Error(
            f"EIA-861 {year}: no Sales-to-Ultimate-Customers rows for utility "
            f"#{utility_number} in {state}"
        )
    return {
        "year": year,
        "utility_number": utility_number,
        "utility_name": util_name or f"#{utility_number}",
        "state": state,
        "ownership": ownership,
        "rows": n_rows,
        "total_sales_mwh": total_sales_mwh,
        "total_customers": total_customers,
        "bundled_revenue_thousand_usd": bundled_rev_thousand,
        "bundled_sales_mwh": bundled_sales_mwh,
    }


def _short_form_name(year: int) -> str:
    return f"Short_Form_{year}.xlsx"


def _reduce_short_form(
    zip_bytes: bytes, *, year: int, utility_number: int, state: str
) -> dict[str, Any]:
    """Reduce the EIA-861 Short Form ("861S") to one utility's annual TOTAL retail line.

    The fallback for a utility absent from the full Sales-to-Ultimate-Customers sheet — a
    municipal or small cooperative that files the short form. The short form carries only
    annual TOTALS (no service-type split): for a full-service municipal that one line *is*
    the bundled retail, so the average price = total revenue / total sales (there are no
    delivery-only rows to exclude). The payload mirrors :func:`_reduce_sales_ult_cust` so the
    downstream model assembly is unchanged, plus a ``form="861s"`` marker for the citations.
    """
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    want = _short_form_name(year).replace(" ", "_").lower()
    matches = [n for n in z.namelist() if n.replace(" ", "_").lower() == want]
    if not matches:
        raise Eia861Error(f"{_short_form_name(year)} not found in EIA-861 {year} zip")
    wb = openpyxl.load_workbook(io.BytesIO(z.read(matches[0])), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # the single "861S" sheet
    out: dict[str, Any] | None = None
    for row in ws.iter_rows(min_row=_SF_DATA_FIRST_ROW, values_only=True):
        if not row:
            continue
        raw_num = row[_SF_COL_UTILNUM - 1]
        if not isinstance(raw_num, (int, float, str)):
            continue
        try:
            num = int(float(raw_num))
        except (TypeError, ValueError):
            continue
        if num != utility_number or str(row[_SF_COL_STATE - 1] or "").strip() != state:
            continue
        sales = _num(row[_SF_COL_TOT_SALES - 1])
        out = {
            "year": year,
            "utility_number": utility_number,
            "utility_name": str(row[_SF_COL_NAME - 1] or "").strip() or f"#{utility_number}",
            "state": state,
            "ownership": str(row[_SF_COL_OWNERSHIP - 1] or "").strip(),
            "rows": 1,
            "total_sales_mwh": sales,
            "total_customers": _num(row[_SF_COL_TOT_CUST - 1]),
            # A full-service municipal/coop's whole retail is bundled — total == bundled.
            "bundled_revenue_thousand_usd": _num(row[_SF_COL_TOT_REV - 1]),
            "bundled_sales_mwh": sales,
            "form": "861s",
        }
        break
    wb.close()
    if out is None:
        raise Eia861Error(
            f"EIA-861 {year}: no Short-Form ('861S') row for utility #{utility_number} in {state}"
        )
    return out


# Curated display names for utilities whose committed reference data predates the connector
# resolving the name itself (keyed by EIA-861 utility number). Any utility not listed uses the
# raw EIA-861 ``utility_name`` from the workbook. #14006 = the AEP-Ohio opco serving Lima +
# Findlay (the curated "AEP Ohio (Ohio Power Company)" label, vs the raw "Ohio Power Co").
_UTILITY_DISPLAY: dict[int, str] = {14006: "AEP Ohio (Ohio Power Company)"}


def fetch_utility_retail(
    *,
    utility_number: int | None = None,
    state: str | None = None,
    year: int | None = None,
    settings: Settings | None = None,
) -> UtilityProfile:
    """The serving utility's EIA-861 retail profile (sales / customers / avg price), cached.

    Defaults to Ohio Power Co (AEP Ohio, #14006) in OH for the configured EIA-861 vintage.
    Totals are summed across service types (Bundled SSO + Delivery shopping); the average
    price is the bundled (full-service) revenue/sales (delivery-only rows exclude generation).
    """
    settings = settings or get_settings()
    utility_number = utility_number or settings.eia861_utility_number
    state = state or settings.eia_state
    year = year or settings.eia861_year
    params = {
        "connector": "eia861",
        "report": "sales_ult_cust",
        "utility_number": utility_number,
        "state": state,
        "year": year,
    }

    def fetch() -> Any:
        zip_bytes = _ensure_zip(settings, year)
        try:
            return _reduce_sales_ult_cust(
                zip_bytes, year=year, utility_number=utility_number, state=state
            )
        except Eia861Error:
            # A small utility (e.g. a municipal) absent from the full Sales sheet files the
            # EIA-861 short form instead — fall back to it before giving up.
            return _reduce_short_form(
                zip_bytes, year=year, utility_number=utility_number, state=state
            )

    payload = cast(
        "dict[str, Any]",
        cached_get(
            "eia861",
            params,
            fetch,
            cache_dir=settings.econ_cache_dir,
            offline=settings.econ_offline,
            fixtures_dir=settings.econ_fixtures_dir,
        ),
    )
    sales_gwh = payload["total_sales_mwh"] / 1000.0
    is_short = payload.get("form") == "861s"
    form_label = "861S Short Form" if is_short else "Sales to Ultimate Customers"
    cite = (
        f"EIA-861 {payload['year']} {form_label}, {payload['utility_name']} "
        f"(#{payload['utility_number']}), {payload['state']}"
    )
    # Average price = bundled (full-service) revenue/sales. Thousand$/MWh == $/kWh, so
    # cents/kWh = (revenue_thousand$ / sales_MWh) * 100. On the full form, delivery-only rows
    # carry only the wires charge (generation paid to a competitive supplier), so a blended
    # price would understate the all-in cost; the short form has no such split (a full-service
    # municipal/coop), so its single total IS the all-in price.
    bundled_sales = payload["bundled_sales_mwh"]
    avg_price = None
    if bundled_sales:
        price_note = (
            "full-service municipal/cooperative retail (short form has no service-type split)"
            if is_short
            else "bundled (full-service) revenue/sales — delivery-only rows exclude "
            "generation, so a blended price understates the all-in cost"
        )
        avg_price = ProvenancedValue.from_connector(
            round(payload["bundled_revenue_thousand_usd"] / bundled_sales * 100.0, 2),
            "cents/kWh",
            citation=f"{cite}; {price_note}",
        )
    total_label = "annual total" if is_short else "bundled + delivery total"
    source_label = "861S short-form total" if is_short else "bundled+delivery total"
    profile = UtilityProfile(
        utility=_UTILITY_DISPLAY.get(utility_number, payload["utility_name"]),
        ownership=str(payload.get("ownership", "")),
        eia_source=f"EIA-861 {payload['year']} per-utility retail (connector; {source_label})",
        retail_sales_gwh=ProvenancedValue.from_connector(
            round(sales_gwh, 1),
            "GWh/yr",
            citation=f"{cite}; {total_label} sales",
        ),
        customers=ProvenancedValue.from_connector(
            round(payload["total_customers"]),
            "customers",
            citation=f"{cite}; {total_label} customers",
        ),
        avg_price_cents_kwh=avg_price,
    )
    log.info(
        "grid.eia861",
        utility=payload["utility_name"],
        sales_gwh=round(sales_gwh),
        customers=round(payload["total_customers"]),
    )
    return profile
